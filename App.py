import tkinter as tk
from tkinter import messagebox, ttk, simpledialog
import pymysql
import openpyxl
import matplotlib.pyplot as plt


class Aplicacion:
    def __init__(self, raiz):
        self.raiz = raiz
        self.raiz.title("Encuesta de Salud y Consumo de Alcohol")
        self.raiz.geometry("800x600")
        self.raiz.configure(bg="#f0f0f0")

        # Menú
        menu_barra = tk.Menu(raiz)
        raiz.config(menu=menu_barra)

        menu_archivo = tk.Menu(menu_barra, tearoff=0)
        menu_barra.add_cascade(label="Archivo", menu=menu_archivo)
        menu_archivo.add_command(label="Salir", command=raiz.quit)

        # Etiquetas y campos de entrada
        campos = [
            ("Edad", "entrada_edad"),
            ("Sexo", "entrada_sexo"),
            ("Bebidas Semanales", "entrada_bebidas"),
            ("Cervezas Semanales", "entrada_cervezas"),
            ("Bebidas Fin de Semana", "entrada_bebidas_fs"),
            ("Bebidas Destiladas Semana", "entrada_bebidas_dest"),
            ("Vinos Semana", "entrada_vinos"),
            ("Pérdidas Control", "entrada_perdidas"),
            ("Diversión Dependencia", "entrada_diversion"),
            ("Problemas Digestivos", "entrada_digestivos"),
            ("Tensión Alta", "entrada_tension"),
            ("Dolor Cabeza", "entrada_dolor")
        ]

        self.entries = {}
        for idx, (texto, var) in enumerate(campos):
            etiqueta = tk.Label(raiz, text=texto + ":", font=("Arial", 12), bg="#f0f0f0")
            etiqueta.grid(row=idx, column=0, sticky="w", padx=10, pady=5)
            entrada = tk.Entry(raiz, font=("Arial", 12))
            entrada.grid(row=idx, column=1, sticky="ew", padx=10, pady=5)
            self.entries[var] = entrada

        # Botones
        self.boton_agregar = tk.Button(raiz, text="Agregar Registro", command=self.agregar_registro, bg="#4CAF50", fg="white", font=("Arial", 12))
        self.boton_agregar.grid(row=len(campos), column=0, sticky="ew", padx=10, pady=5)

        self.boton_ver = tk.Button(raiz, text="Ver Registros", command=self.ver_registros, bg="#2196F3", fg="white", font=("Arial", 12))
        self.boton_ver.grid(row=len(campos), column=1, sticky="ew", padx=10, pady=5)

        self.boton_actualizar = tk.Button(raiz, text="Actualizar Registro", command=self.actualizar_registro, bg="#FFC107", fg="white", font=("Arial", 12))
        self.boton_actualizar.grid(row=len(campos) + 1, column=0, sticky="ew", padx=10, pady=5)

        self.boton_eliminar = tk.Button(raiz, text="Eliminar Registro", command=self.eliminar_registro, bg="#F44336", fg="white", font=("Arial", 12))
        self.boton_eliminar.grid(row=len(campos) + 1, column=1, sticky="ew", padx=10, pady=5)

        self.boton_excel = tk.Button(raiz, text="Crear/Actualizar Excel", command=self.crear_actualizar_excel, bg="#9C27B0", fg="white", font=("Arial", 12))
        self.boton_excel.grid(row=len(campos) + 2, column=0, columnspan=2, sticky="ew", padx=10, pady=5)

        self.boton_graficos = tk.Button(raiz, text="Mostrar Gráficos", command=self.mostrar_graficos, bg="#FF5722", fg="white", font=("Arial", 12))
        self.boton_graficos.grid(row=len(campos) + 3, column=0, columnspan=2, sticky="ew", padx=10, pady=5)

        # Tabla de registros con scroll
        self.tree = ttk.Treeview(raiz,
                                 columns=("ID", "edad", "sexo", "bebidas", "cervezas", "bebidas_fs", "bebidas_dest",
                                          "vinos", "perdidas", "diversion", "digestivos", "tension", "dolor"),
                                 show="headings")
        for col in self.tree["columns"]:
            self.tree.heading(col, text=col.capitalize())
        self.tree.grid(row=len(campos) + 4, column=0, columnspan=2, sticky="nsew", padx=10, pady=5)

        # Scrollbar vertical
        scrollbar_vertical = ttk.Scrollbar(raiz, orient="vertical", command=self.tree.yview)
        scrollbar_vertical.grid(row=len(campos) + 4, column=2, sticky="ns")
        self.tree.configure(yscrollcommand=scrollbar_vertical.set)

        # Scrollbar horizontal
        scrollbar_horizontal = ttk.Scrollbar(raiz, orient="horizontal", command=self.tree.xview)
        scrollbar_horizontal.grid(row=len(campos) + 5, column=0, columnspan=2, sticky="ew")
        self.tree.configure(xscrollcommand=scrollbar_horizontal.set)

        # Hacer la ventana responsive
        raiz.grid_rowconfigure(len(campos), weight=1)
        raiz.grid_columnconfigure(0, weight=1)
        raiz.grid_columnconfigure(1, weight=1)

        # Vincular el evento de seleccionar un registro en la tabla
        self.tree.bind("<ButtonRelease-1>", self.seleccionar_registro)

    def obtener_campos(self):
        return {
            "edad": self.entries["entrada_edad"].get(),
            "sexo": self.entries["entrada_sexo"].get(),
            "bebidas": self.entries["entrada_bebidas"].get(),
            "cervezas": self.entries["entrada_cervezas"].get(),
            "bebidas_fs": self.entries["entrada_bebidas_fs"].get(),
            "bebidas_dest": self.entries["entrada_bebidas_dest"].get(),
            "vinos": self.entries["entrada_vinos"].get(),
            "perdidas": self.entries["entrada_perdidas"].get(),
            "diversion": self.entries["entrada_diversion"].get(),
            "digestivos": self.entries["entrada_digestivos"].get(),
            "tension": self.entries["entrada_tension"].get(),
            "dolor": self.entries["entrada_dolor"].get()
        }

    def agregar_registro(self):
        datos = self.obtener_campos()
        try:
            nuevo_id = obtener_siguiente_id()
            agregar_registro_db(
                id=nuevo_id,
                edad=datos["edad"],
                sexo=datos["sexo"],
                bebidas=datos["bebidas"],
                cervezas=datos["cervezas"],
                bebidas_fs=datos["bebidas_fs"],
                bebidas_dest=datos["bebidas_dest"],
                vinos=datos["vinos"],
                perdidas=datos["perdidas"],
                diversion=datos["diversion"],
                digestivos=datos["digestivos"],
                tension=datos["tension"],
                dolor=datos["dolor"]
            )
            messagebox.showinfo("Éxito", "Registro agregado exitosamente")
            self.ver_registros()
        except Exception as e:
            messagebox.showerror("Error", f"Error al agregar el registro: {e}")
            print(f"Error al agregar el registro: {e}")

    def ver_registros(self):
        try:
            for row in self.tree.get_children():
                self.tree.delete(row)

            registros = obtener_registros_db()
            if not registros:
                messagebox.showwarning("Sin Registros", "No se encontraron registros en la base de datos.")
                return

            for registro in registros:
                self.tree.insert("", "end", values=registro)
        except Exception as e:
            messagebox.showerror("Error", f"Error al ver los registros: {e}")
            print(f"Error al ver los registros: {e}")

    def seleccionar_registro(self, event):
        selected_item = self.tree.selection()
        if selected_item:
            record_id = self.tree.item(selected_item[0])['values'][0]
            registro = obtener_registro_por_id(record_id)

            # Rellenar los campos con los valores del registro
            self.entries["entrada_edad"].delete(0, tk.END)
            self.entries["entrada_edad"].insert(0, registro["edad"])

            self.entries["entrada_sexo"].delete(0, tk.END)
            self.entries["entrada_sexo"].insert(0, registro["sexo"])

            self.entries["entrada_bebidas"].delete(0, tk.END)
            self.entries["entrada_bebidas"].insert(0, registro["bebidas"])

            self.entries["entrada_cervezas"].delete(0, tk.END)
            self.entries["entrada_cervezas"].insert(0, registro["cervezas"])

            self.entries["entrada_bebidas_fs"].delete(0, tk.END)
            self.entries["entrada_bebidas_fs"].insert(0, registro["bebidas_fs"])

            self.entries["entrada_bebidas_dest"].delete(0, tk.END)
            self.entries["entrada_bebidas_dest"].insert(0, registro["bebidas_dest"])

            self.entries["entrada_vinos"].delete(0, tk.END)
            self.entries["entrada_vinos"].insert(0, registro["vinos"])

            self.entries["entrada_perdidas"].delete(0, tk.END)
            self.entries["entrada_perdidas"].insert(0, registro["perdidas"])

            self.entries["entrada_diversion"].delete(0, tk.END)
            self.entries["entrada_diversion"].insert(0, registro["diversion"])

            self.entries["entrada_digestivos"].delete(0, tk.END)
            self.entries["entrada_digestivos"].insert(0, registro["digestivos"])

            self.entries["entrada_tension"].delete(0, tk.END)
            self.entries["entrada_tension"].insert(0, registro["tension"])

            self.entries["entrada_dolor"].delete(0, tk.END)
            self.entries["entrada_dolor"].insert(0, registro["dolor"])

    def actualizar_registro(self):
        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showwarning("Selección inválida", "Seleccione un registro para actualizar.")
            return

        record_id = self.tree.item(selected_item[0])['values'][0]
        datos = self.obtener_campos()

        try:
            # Llamar a la función para actualizar el registro en la base de datos
            actualizar_registro_db(record_id, **datos)
            messagebox.showinfo("Éxito", "Registro actualizado exitosamente")

            # Limpiar los campos de entrada después de la actualización
            for entry in self.entries.values():
                entry.delete(0, tk.END)

            self.ver_registros()  # Refrescar la vista con los registros actualizados
        except Exception as e:
            messagebox.showerror("Error", f"Error al actualizar el registro: {e}")
            print(f"Error al actualizar el registro: {e}")

    def eliminar_registro(self):
        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showwarning("Selección inválida", "Seleccione un registro para eliminar.")
            return

        record_id = self.tree.item(selected_item[0])['values'][0]
        try:
            eliminar_registro_db(record_id)
            messagebox.showinfo("Éxito", "Registro eliminado exitosamente")
            self.ver_registros()  # Refrescar la vista con los registros actualizados
        except Exception as e:
            messagebox.showerror("Error", f"Error al eliminar el registro: {e}")
            print(f"Error al eliminar el registro: {e}")

    def crear_actualizar_excel(self):
        try:
            # Solicitar criterios de filtro
            filtro_edad = simpledialog.askstring("Filtro", "Ingrese la edad máxima (o deje en blanco para todos):")
            filtro_sexo = simpledialog.askstring("Filtro", "Ingrese el sexo (o deje en blanco para todos):")
            filtro_bebidas = simpledialog.askstring("Filtro",
                                                    "Ingrese el número máximo de bebidas semanales (o deje en blanco para todos):")

            registros = obtener_registros_db()
            if not registros:
                messagebox.showwarning("Sin Registros", "No se encontraron registros en la base de datos.")
                return

            # Filtrar registros
            registros_filtrados = []
            for registro in registros:
                edad_valida = not filtro_edad or (registro[1] is not None and int(registro[1]) <= int(filtro_edad))
                sexo_valido = not filtro_sexo or (
                            registro[2] is not None and registro[2].lower() == filtro_sexo.lower())
                bebidas_validas = not filtro_bebidas or (
                            registro[3] is not None and int(registro[3]) <= int(filtro_bebidas))

                if edad_valida and sexo_valido and bebidas_validas:
                    registros_filtrados.append(registro)

            if not registros_filtrados:
                messagebox.showwarning("Sin Registros", "No se encontraron registros que coincidan con los filtros.")
                return

            # Crear o cargar el archivo Excel
            try:
                wb = openpyxl.load_workbook("registros.xlsx")
                ws = wb.active
            except FileNotFoundError:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.append(["ID", "Edad", "Sexo", "Bebidas Semanales", "Cervezas Semanales", "Bebidas Fin de Semana",
                           "Bebidas Destiladas Semana", "Vinos Semana", "Pérdidas Control", "Diversión Dependencia",
                           "Problemas Digestivos", "Tensión Alta", "Dolor Cabeza"])

            # Limpiar las filas existentes
            ws.delete_rows(2, ws.max_row)

            # Añadir los registros filtrados
            for registro in registros_filtrados:
                ws.append(registro)

            # Guardar el archivo
            wb.save("registros.xlsx")
            messagebox.showinfo("Éxito", "Archivo Excel creado/actualizado exitosamente")
        except Exception as e:
            messagebox.showerror("Error", f"Error al crear/actualizar el archivo Excel: {e}")
            print(f"Error al crear/actualizar el archivo Excel: {e}")

    def mostrar_graficos(self):
        try:
            registros = obtener_registros_db()
            if not registros:
                messagebox.showwarning("Sin Registros", "No se encontraron registros en la base de datos.")
                return

            edades = [registro[1] for registro in registros]
            bebidas = [registro[3] for registro in registros]

            plt.figure(figsize=(10, 5))
            plt.scatter(edades, bebidas, c='blue', label='Bebidas Semanales')
            plt.xlabel('Edad')
            plt.ylabel('Bebidas Semanales')
            plt.title('Consumo de Bebidas Semanales por Edad')
            plt.legend()
            plt.grid(True)
            plt.show()
        except Exception as e:
            messagebox.showerror("Error", f"Error al mostrar los gráficos: {e}")
            print(f"Error al mostrar los gráficos: {e}")

# Funciones para interactuar con la base de datos

def obtener_conexion_db():
    return pymysql.connect(host="localhost", user="root", password="curso", database="encuestas")

def obtener_siguiente_id():
    try:
        conn = obtener_conexion_db()
        cursor = conn.cursor()
        cursor.execute("SELECT MAX(idEncuesta) FROM ENCUESTA")
        max_id = cursor.fetchone()[0]
        return max_id + 1 if max_id else 1
    except Exception as e:
        print(f"Error al obtener el siguiente ID: {e}")
        return 1
    finally:
        conn.close()

def agregar_registro_db(id, edad, sexo, bebidas, cervezas, bebidas_fs, bebidas_dest, vinos, perdidas, diversion,
                         digestivos, tension, dolor):
    try:
        conn = obtener_conexion_db()
        cursor = conn.cursor()
        query = """
            INSERT INTO ENCUESTA (idEncuesta, edad, Sexo, BebidasSemana, CervezasSemana, BebidasFinSemana,
                                  BebidasDestiladasSemana, VinosSemana, PerdidasControl, DiversionDependenciaAlcohol,
                                  ProblemasDigestivos, TensionAlta, DolorCabeza)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        cursor.execute(query, (id, edad, sexo, bebidas, cervezas, bebidas_fs, bebidas_dest, vinos, perdidas, diversion,
                               digestivos, tension, dolor))
        conn.commit()
    except Exception as e:
        print(f"Error al agregar el registro: {e}")
    finally:
        conn.close()

def obtener_registros_db():
    try:
        conn = obtener_conexion_db()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM ENCUESTA")
        registros = cursor.fetchall()
        return registros
    except Exception as e:
        print(f"Error al obtener los registros: {e}")
        return []
    finally:
        conn.close()

def obtener_registro_por_id(id):
    try:
        conn = obtener_conexion_db()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM ENCUESTA WHERE idEncuesta = %s", (id,))
        registro = cursor.fetchone()
        if registro:
            return {
                "edad": registro[1],
                "sexo": registro[2],
                "bebidas": registro[3],
                "cervezas": registro[4],
                "bebidas_fs": registro[5],
                "bebidas_dest": registro[6],
                "vinos": registro[7],
                "perdidas": registro[8],
                "diversion": registro[9],
                "digestivos": registro[10],
                "tension": registro[11],
                "dolor": registro[12]
            }
        return None
    except Exception as e:
        print(f"Error al obtener el registro por ID: {e}")
        return None
    finally:
        conn.close()

def actualizar_registro_db(id, edad, sexo, bebidas, cervezas, bebidas_fs, bebidas_dest, vinos, perdidas, diversion,
                            digestivos, tension, dolor):
    try:
        conn = obtener_conexion_db()
        cursor = conn.cursor()
        query = """
            UPDATE ENCUESTA SET edad = %s, Sexo = %s, BebidasSemana = %s, CervezasSemana = %s, BebidasFinSemana = %s,
                BebidasDestiladasSemana = %s, VinosSemana = %s, PerdidasControl = %s, DiversionDependenciaAlcohol = %s,
                Problemasdigestivos = %s, TensionAlta = %s, DolorCabeza = %s
            WHERE idEncuesta = %s
        """
        cursor.execute(query, (edad, sexo, bebidas, cervezas, bebidas_fs, bebidas_dest, vinos, perdidas, diversion,
                               digestivos, tension, dolor, id))
        conn.commit()
    except Exception as e:
        print(f"Error al actualizar el registro: {e}")
    finally:
        conn.close()

def eliminar_registro_db(id):
    try:
        conn = obtener_conexion_db()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM ENCUESTA WHERE idEncuesta = %s", (id,))
        conn.commit()
    except Exception as e:
        print(f"Error al eliminar el registro: {e}")
    finally:
        conn.close()

# Crear ventana principal y ejecutar la aplicación
if __name__ == "__main__":
    raiz = tk.Tk()
    app = Aplicacion(raiz)
    raiz.mainloop()